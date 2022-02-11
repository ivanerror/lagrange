from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import math
import matplotlib.pyplot as plt
import plotly.graph_objects as go

# Create Workbook Instance
wb = Workbook()
# load existing workbook 
wb = load_workbook('lagrange.xlsx')

ws = wb.active

row_1 = ws['1']
row_2 = ws['2']

xTab = []
fTab = []
yList = []
eList = []
nList = []

for cell in row_1:
    xTab.append(cell.value)

for cell in row_2:
    fTab.append(cell.value)

zipped_lists = zip(xTab, fTab)
sorted_pairs = sorted(zipped_lists)

tuples = zip(*sorted_pairs)
xTab, fTab = [list(tuple) for tuple in tuples]

print(xTab)
print(fTab)

checkY = True
while(checkY):
    y = float(input('x = '))

    if y < xTab[0]:
        print("Data terlalu Kecil")
    elif y > xTab[len(xTab)-1]:
        print("Data terlalu Besar")
    else:
        checkY = False

ytru = ((math.exp(1)**((3/2)*y+1))/2)-3*math.sqrt(y)
inp = int(input('N = '))

for n in range(inp):
    take = n+1

    for index in range(len(xTab)):
        if xTab[index] >= y:
            if y-xTab[index-1] > xTab[index]-y:
                left = (index-1)
                right = index+take
                if index+take > len(xTab):
                    right = len(xTab)
                    left = left-(index+take - len(xTab))
                x = xTab[left:right]
                f = fTab[left:right]
                break
            else:
                left = index-take
                right = index+1
                if(index-take < 0):
                    left = 0
                    right = right+(take-index)
                x = xTab[left:right]
                f = fTab[left:right]
                break

    total = 0
    #print(x)
    #print(f)

    for i in range(take):
        L = 1
        j = 0
        for j in range(take):
            if j != i:
                L = L*((y-x[j]) / (x[i]-x[j]))
        #print('l',i+1, L)
        total = total + f[i] * L
    print('x :', y, '; f(x) :', total)
    ea = abs((ytru-total)/ytru)*100
    print('est. error :', ea)
    yList.append(total)
    eList.append(ea)
    nList.append(i+1)

fig = go.Figure(data=[go.Table(header=dict(values=['Orde ke-', 'Nilai f(x)', 'Error Rate (%)']),
                 cells=dict(values=[nList, yList, eList]))
                     ])
fig.update_layout(width=1000)
fig.show()

plt.subplot(2, 1, 1)
plt.plot(nList, yList, color='blue', marker='o', linewidth=0.7, label="Nilai fungsi")
plt.axhline(y=ytru, color='r', linestyle='--', linewidth=0.7, label="Nilai sebenarnya")
plt.title('Estimasi Nilai Fungsi')
plt.xlabel('Orde ke-N')
plt.ylabel('f(x)')
plt.grid(True)
plt.legend()

plt.subplot(2, 1, 2)
plt.plot(nList, eList, color='red', marker='o', linewidth=0.7, label="Error Rate")
plt.axhline(y=0, color='green', linestyle='--', linewidth=0.7, label="Kesalahan 0%")
plt.title('Estimasi Kesalahan (%)')
plt.xlabel('Orde ke-N')
plt.ylabel('f(x)')
plt.grid(True)

plt.tight_layout()
plt.legend()
plt.show()