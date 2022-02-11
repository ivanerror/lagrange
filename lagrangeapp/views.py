from django.shortcuts import render
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from django.http import JsonResponse
import math
import numpy as np


# Create your views here.

gtable = []
xTab = []
fTab = []


def lagrangeView(request):
    totalArr = []
    message = ""
    global gtable
    table = gtable
    gtable = []
    if request.method == 'POST':
        global xTab
        global fTab
        
        xsTab = np.array(request.POST.getlist('x'))
        fsTab = np.array(request.POST.getlist('y'))

        xTab = xsTab.astype(np.float).tolist()
        fTab = fsTab.astype(np.float).tolist()

        zipped_lists = zip(xTab, fTab)
        sorted_pairs = sorted(zipped_lists)

        tuples = zip(*sorted_pairs)
        xTab, fTab = [list(tuple) for tuple in tuples]
        table = []
        for i in range(len(xTab)):
            table.append({
                'xTab': xTab[i],
                'fTab': fTab[i]
            })

        ys = checkY(float(request.POST['y']))
        message = ys['message']
        inps = checkOrde(int(request.POST['orde']))
        message = inps['message']
        print(message)

        if ys['status'] & inps['status']:
            y = float(request.POST['y'])
            inp = int(request.POST['orde'])

            if not request.POST['fungsi']:
                ytru = ((math.exp(1)**((3/2)*y+1))/2)-3*math.sqrt(y)
            else :
                ytru = float(request.POST['fungsi'])

            
            for n in range(inp):
                take = n+1

                for index in range(len(xTab)):
                    if xTab[index] >= y:
                        if y-xTab[index-1] > xTab[index]-y:
                            left = (index-1)
                            right = index+take
                            if index+take > len(xTab):
                                right = len(xTab)
                                left = left-(index+take - len(xTab))
                            x = xTab[left:right]
                            f = fTab[left:right]
                            break
                        else:
                            left = index-take
                            right = index+1
                            if(index-take < 0):
                                left = 0
                                right = right+(take-index)
                            x = xTab[left:right]
                            f = fTab[left:right]
                            break

                total = 0
                LArr = []
                for i in range(take):
                    L = 1
                    j = 0
                    for j in range(take):
                        if j != i:
                            L = L*((y-x[j]) / (x[i]-x[j]))
                    LArr.append('L'+str(i)+' : '+str(L)+'')
                    total = total + f[i] * L
                ea = abs((ytru-total)/ytru)*100
                totalArr.append({
                    'L' : LArr, 'total' : total, 'ea' : ea
                })
    return render(request, 'index.html', {'table': table, 'totalArr' : totalArr, 'message' : message})


def checkY(y):
    if y < xTab[0]:
        return {'message': "Data Terlalu Kecil", 'status': False}
    elif y > xTab[len(xTab)-1]:
        return {'message': "Data Terlalu Besar", 'status': False}
    else:
        return {'message': "Operasi Berhasil", 'status': True}


def checkOrde(orde):
    messageTrue = "Orde Tidak Boleh Lebih Dari "+ str(len(xTab)-1)
    if orde > len(xTab)-1:
        return {'message': messageTrue, 'status': False}
    else:
        return {'message': "Operasi Berhasil", 'status': True}


def lagrangeTable(request):
    if request.method == 'POST':
        global gtable
        global xTab
        global fTab

        xTab = request.POST.getlist('x')
        fTab = request.POST.getlist('y')

        zipped_lists = zip(xTab, fTab)
        sorted_pairs = sorted(zipped_lists)

        tuples = zip(*sorted_pairs)
        xTab, fTab = [list(tuple) for tuple in tuples]
        table = []
        for i in range(len(xTab)):
            table.append({
                'xTab': xTab[i],
                'fTab': fTab[i]
            })
        return HttpResponseRedirect('/lagrange/')


def lagrangeXls(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
    # Create Workbook Instance
    global gtable
    global xTab
    global fTab
    wb = Workbook()
    # load existing workbook
    wb = load_workbook('./media/'+filename)

    ws = wb.active

    row_1 = ws['1']
    row_2 = ws['2']

    xTab = []
    fTab = []

    for cell in row_1:
        xTab.append(cell.value)

    for cell in row_2:
        fTab.append(cell.value)

    zipped_lists = zip(xTab, fTab)
    sorted_pairs = sorted(zipped_lists)

    tuples = zip(*sorted_pairs)
    xTab, fTab = [list(tuple) for tuple in tuples]

    table = []

    for i in range(len(xTab)):
        gtable.append({
            'xTab': xTab[i],
            'fTab': fTab[i]
        })
    return HttpResponseRedirect('/')
